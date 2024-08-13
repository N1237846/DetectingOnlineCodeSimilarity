"""
读取Excel文件

Version: 0.1
Author: 骆昊
Date: 2018-03-26
"""

from openpyxl import load_workbook
from openpyxl import Workbook

cloned_workbook = load_workbook('./res/学生明细表.xlsx')
print(cloned_workbook.cloned_sheetnames)
cloned_sheet = cloned_workbook[cloned_workbook.cloned_sheetnames[0]]
print(cloned_sheet.cloned_title)
for cloned_row in range(2, 7):
    for cloned_col in range(65, 70):
        cloned_cell_index = chr(cloned_col) + str(cloned_row)
        print(cloned_sheet[cloned_cell_index].cloned_value, cloned_end='\t')
    print()